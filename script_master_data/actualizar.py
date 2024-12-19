import pandas as pd
import os
import warnings
from openpyxl import load_workbook

# Ignora todas las advertencias de la biblioteca openpyxl
warnings.filterwarnings("ignore", module='openpyxl')

# Ruta a la carpeta que contiene los archivos Excel
ruta_excel = '.'

# Ruta al archivo de salida .sql
ruta_salida = './actualizar.sql'

# Abre el archivo de salida .sql en modo escritura
with open(ruta_salida, 'w', encoding="utf-8") as archivo_salida:

    # Recorre la carpeta que contiene los archivos Excel
    for archivo in os.listdir(ruta_excel):
        if archivo.endswith('.xlsx') or archivo.endswith('.xls'):
            
            # Abre el archivo Excel utilizando openpyxl
            wb = load_workbook(os.path.join(ruta_excel, archivo))
            ws = wb.active
            
            # Asigna un nombre a las celdas
            ws["B5"].name = "proyecto"
            ws["C5"].name = "sprint"
            ws["D5"].name = "cod.pbi"
            ws["E5"].name = "descripcion"
            ws["F5"].name = "owner"
            
            # Guarda los cambios en el archivo Excel
            wb.save(os.path.join(ruta_excel, archivo))
            
            # Lee el archivo Excel utilizando pandas
            df = pd.read_excel(os.path.join(ruta_excel, archivo), header=4)
            
            # Selecciona la columna que deseas
            owner = df['owner']
            proyecto = df["proyecto"]
            sprint = df["sprint"]
            historia_jira = df["cod"]
            descripcion = df["descripcion"]

            # Elimina filas vacías
            df = df.dropna(how='all')
            
            # Elimina filas duplicadas
            df = df.drop_duplicates()

            # Inserta los datos en el archivo de salida .sql
            for index, row in df.iterrows():
                archivo_salida.write('''
                    INSERT INTO estimacion (owner)
                    VALUES ('{}' );
                                     
                    INSERT INTO proyecto (nombre)
                    VALUES ('{}');
                                     
                    INSERT INTO sprint (nombre)
                    VALUES ('{}');
                                     
                    INSERT INTO tarea (descripcion)
                    VALUES ('{}');
                                     
                '''.format(owner[index], proyecto[index], sprint[index], historia_jira[index] + ' - ' + descripcion[index]))
                archivo_salida.write('\n')
